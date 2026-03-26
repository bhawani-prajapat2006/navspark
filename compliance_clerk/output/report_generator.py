"""
Excel/CSV report generator for extracted document data.

Converts consolidated extraction results into styled Excel
spreadsheets or CSV files matching the expected output format.
"""

import csv
import logging
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from compliance_clerk.models.schemas import ConsolidatedRow
from compliance_clerk.config import DEFAULT_OUTPUT_PATH

logger = logging.getLogger(__name__)

# Excel header styling
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column widths for readability
COLUMN_WIDTHS = {
    "A": 8,   # Sr.no.
    "B": 16,  # Village
    "C": 14,  # Survey No.
    "D": 18,  # Area in NA Order
    "E": 14,  # Dated
    "F": 28,  # NA Order No.
    "G": 20,  # Lease Deed Doc. No.
    "H": 14,  # Lease Area
    "I": 14,  # Lease Start
}


def generate_excel(
    rows: list[ConsolidatedRow],
    output_path: Optional[str | Path] = None,
) -> Path:
    """
    Generate a styled Excel report from consolidated extraction data.

    Args:
        rows: List of ConsolidatedRow objects.
        output_path: Path for the output Excel file.

    Returns:
        Path to the generated Excel file.
    """
    output_path = Path(output_path or DEFAULT_OUTPUT_PATH)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write headers
    headers = [
        "Sr.no.", "Village", "Survey No.", "Area in NA Order",
        "Dated", "NA Order No.", "Lease Deed Doc. No.",
        "Lease Area", "Lease Start",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        data = row.to_excel_dict()
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=data[header])
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    # Set column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Save
    wb.save(str(output_path))
    logger.info(f"Excel report saved: {output_path} ({len(rows)} rows)")

    return output_path


def generate_csv(
    rows: list[ConsolidatedRow],
    output_path: Optional[str | Path] = None,
) -> Path:
    """
    Generate a CSV report from consolidated extraction data.

    Args:
        rows: List of ConsolidatedRow objects.
        output_path: Path for the output CSV file.

    Returns:
        Path to the generated CSV file.
    """
    if output_path is None:
        output_path = DEFAULT_OUTPUT_PATH.with_suffix(".csv")
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    headers = [
        "Sr.no.", "Village", "Survey No.", "Area in NA Order",
        "Dated", "NA Order No.", "Lease Deed Doc. No.",
        "Lease Area", "Lease Start",
    ]

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.to_excel_dict())

    logger.info(f"CSV report saved: {output_path} ({len(rows)} rows)")
    return output_path


if __name__ == "__main__":
    """Quick test with sample data."""
    sample_rows = [
        ConsolidatedRow(
            sr_no=1, village="Rampura Mota", survey_number="257",
            area_in_na_order=16534, dated="7/1/2026",
            na_order_number="iORA/31/02/112/7/2026",
            lease_deed_doc_number="837/2025", lease_area=16792,
            lease_start="28/05/2025",
        ),
        ConsolidatedRow(
            sr_no=2, village="Rampura Mota", survey_number="251/p2",
            area_in_na_order=5997, dated="16/02/2026",
            na_order_number="iORA/31/02/112/25/2026",
            lease_deed_doc_number="1141/2026", lease_area=5997,
            lease_start="21/01/2026",
        ),
    ]

    # Test Excel
    xlsx_path = generate_excel(sample_rows, "/tmp/test_output.xlsx")
    print(f"Excel created: {xlsx_path} ✓")

    # Test CSV
    csv_path = generate_csv(sample_rows, "/tmp/test_output.csv")
    print(f"CSV created: {csv_path} ✓")

    print("\nReport generator tests passed!")
